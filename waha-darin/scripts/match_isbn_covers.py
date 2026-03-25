from __future__ import annotations

import argparse
import csv
import os
import re
from pathlib import Path
from typing import Any, Iterable, Optional


ISBN_COL_NAME_RE = re.compile(r"^\s*isbn(\s*13)?\s*$", re.IGNORECASE)
ISBN_VALUE_RE = re.compile(r"^[0-9X]{10,13}$", re.IGNORECASE)  # ISBN10/13


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Match XLSX ISBNs to local cover filenames")
    p.add_argument("--xlsx", required=True, help="Path to Excel file")
    p.add_argument(
        "--covers-dir",
        default=str(Path(__file__).resolve().parents[1] / "storage" / "app" / "covers_by_isbn"),
        help="Directory containing exported covers named by ISBN",
    )
    p.add_argument("--sheet", default=None, help="Sheet name (default: active sheet)")
    p.add_argument("--max-scan-rows", type=int, default=300, help="Rows to scan when guessing ISBN column")
    p.add_argument("--write-report", action="store_true", help="Write CSV report next to covers dir")
    return p.parse_args()


def norm_isbn(val: Any) -> str:
    if val is None:
        return ""
    # Excel often stores ISBN as a number (float). Handle that safely.
    if isinstance(val, bool):
        return ""
    if isinstance(val, int):
        s = str(val)
        return s if s and s != "0" else ""
    if isinstance(val, float):
        if val.is_integer():
            s = str(int(val))
            return s if s and s != "0" else ""
        # non-integer floats are not valid ISBNs
        return ""
    s = str(val).strip()
    if not s or s == "0":
        return ""
    # common case: "978... .0" coming from csv-like conversions
    if s.endswith(".0") and s[:-2].replace(" ", "").isdigit():
        s = s[:-2]
    # remove common separators/spaces
    s = re.sub(r"[^0-9Xx]+", "", s)
    s = s.upper()
    if s in {"", "0"}:
        return ""
    return s


def is_isbn_like(s: str) -> bool:
    if not s:
        return False
    if not ISBN_VALUE_RE.match(s):
        return False
    # reject too-short/too-long extreme noise
    return 10 <= len(s) <= 17


def iter_rows(ws, start_row: int = 1) -> Iterable[list[Any]]:
    for row in ws.iter_rows(min_row=start_row, values_only=True):
        yield list(row)


def find_header_row_and_isbn_col(ws, max_header_scan_rows: int = 30) -> tuple[int, Optional[int], str]:
    """
    Returns (header_row_idx_1based, isbn_col_idx_0based_or_None, method)
    """
    # 1) Look for a header cell that is literally "ISBN" (or similar)
    for r_idx in range(1, max_header_scan_rows + 1):
        row = next(iter_rows(ws, start_row=r_idx), None)
        if not row:
            continue
        for c_idx, cell in enumerate(row):
            name = str(cell).strip() if cell is not None else ""
            if name and ISBN_COL_NAME_RE.match(name):
                return (r_idx, c_idx, "header_name_match")

    # 2) Assume first row is header; try to locate isbn column by name anyway
    row1 = next(iter_rows(ws, start_row=1), []) or []
    for c_idx, cell in enumerate(row1):
        name = str(cell).strip() if cell is not None else ""
        if name and ISBN_COL_NAME_RE.match(name):
            return (1, c_idx, "row1_header_name_match")

    # 3) Guess by data density: pick column with most isbn-like values
    # Assume header is row 1; data starts at row 2 (we may adjust below).
    best_col = None
    best_score = -1
    sample_rows = []
    for i, row in enumerate(iter_rows(ws, start_row=2), start=0):
        sample_rows.append(row)
        if i + 1 >= 300:
            break

    if not sample_rows:
        return (1, None, "no_data")

    max_cols = max(len(r) for r in sample_rows)
    for c_idx in range(max_cols):
        score = 0
        for row in sample_rows:
            if c_idx >= len(row):
                continue
            v = norm_isbn(row[c_idx])
            if is_isbn_like(v):
                score += 1
        if score > best_score:
            best_score = score
            best_col = c_idx

    # require a minimum to avoid picking random numeric column
    if best_score >= 10:
        # Heuristic: if row 1, that column looks like an ISBN value, row 1 is data not header.
        row1 = next(iter_rows(ws, start_row=1), []) or []
        cell1 = row1[best_col] if best_col is not None and best_col < len(row1) else None
        cell1n = norm_isbn(cell1)
        if is_isbn_like(cell1n):
            return (0, best_col, f"guessed_by_density_no_header(score={best_score})")
        return (1, best_col, f"guessed_by_density(score={best_score})")

    return (1, None, f"no_isbn_col_detected(best_score={best_score})")


def scan_xlsx_isbns(xlsx_path: Path, sheet: Optional[str], max_scan_rows: int) -> tuple[set[str], dict[str, Any]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as e:  # pragma: no cover
        raise SystemExit(f"openpyxl is required to read XLSX. Error: {e}")

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb.active

    header_row, isbn_col, method = find_header_row_and_isbn_col(ws)
    meta: dict[str, Any] = {
        "sheet": ws.title,
        "header_row": header_row,
        "isbn_col": isbn_col,
        "detection_method": method,
    }
    if isbn_col is None:
        return set(), meta

    isbns: set[str] = set()
    total_rows = 0
    empty_isbn = 0
    bad_isbn = 0

    start_row = 1 if header_row == 0 else (header_row + 1)
    for i, row in enumerate(iter_rows(ws, start_row=start_row), start=1):
        if max_scan_rows > 0 and i > max_scan_rows:
            break
        total_rows += 1
        v = row[isbn_col] if isbn_col < len(row) else None
        n = norm_isbn(v)
        if not n:
            empty_isbn += 1
            continue
        if not is_isbn_like(n):
            bad_isbn += 1
            continue
        isbns.add(n)

    meta.update(
        {
            "scanned_rows": total_rows,
            "empty_isbn_rows": empty_isbn,
            "bad_isbn_rows": bad_isbn,
            "unique_isbns": len(isbns),
        }
    )
    return isbns, meta


def scan_cover_isbns(covers_dir: Path) -> tuple[set[str], dict[str, Any]]:
    if not covers_dir.exists() or not covers_dir.is_dir():
        raise SystemExit(f"covers dir not found: {covers_dir}")

    cover_isbns: set[str] = set()
    total_files = 0
    isbn_named_files = 0

    for p in covers_dir.iterdir():
        if not p.is_file():
            continue
        if p.name.startswith("_"):
            continue
        total_files += 1
        base = p.stem
        n = norm_isbn(base)
        if is_isbn_like(n) and n == base.upper():
            isbn_named_files += 1
            cover_isbns.add(n)

    meta = {
        "covers_dir": str(covers_dir),
        "total_files": total_files,
        "isbn_named_files": isbn_named_files,
        "unique_cover_isbns": len(cover_isbns),
    }
    return cover_isbns, meta


def main() -> int:
    args = parse_args()
    xlsx = Path(args.xlsx)
    covers_dir = Path(args.covers_dir)

    excel_isbns, excel_meta = scan_xlsx_isbns(xlsx, args.sheet, args.max_scan_rows)
    cover_isbns, cover_meta = scan_cover_isbns(covers_dir)

    matched = excel_isbns & cover_isbns
    excel_only = excel_isbns - cover_isbns
    covers_only = cover_isbns - excel_isbns

    print("XLSX:")
    print(f"  file: {xlsx}")
    print(f"  sheet: {excel_meta.get('sheet')}")
    print(f"  header_row: {excel_meta.get('header_row')}  isbn_col: {excel_meta.get('isbn_col')}")
    print(f"  detection: {excel_meta.get('detection_method')}")
    print(f"  scanned_rows: {excel_meta.get('scanned_rows')}")
    print(f"  unique_isbns: {excel_meta.get('unique_isbns')}")
    print("")
    print("COVERS:")
    print(f"  dir: {covers_dir}")
    print(f"  total_files: {cover_meta.get('total_files')}")
    print(f"  isbn_named_files: {cover_meta.get('isbn_named_files')}")
    print(f"  unique_cover_isbns: {cover_meta.get('unique_cover_isbns')}")
    print("")
    print("MATCHING (by ISBN):")
    print(f"  matched_unique_isbns: {len(matched)}")
    print(f"  excel_isbns_without_cover: {len(excel_only)}")
    print(f"  cover_isbns_not_in_excel: {len(covers_only)}")

    if args.write_report:
        report_path = covers_dir / "_isbn_match_report.csv"
        with report_path.open("w", encoding="utf-8", newline="") as f:
            w = csv.writer(f)
            w.writerow(["isbn", "in_excel", "has_cover"])
            for isbn in sorted(matched):
                w.writerow([isbn, 1, 1])
            for isbn in sorted(excel_only):
                w.writerow([isbn, 1, 0])
            for isbn in sorted(covers_only):
                w.writerow([isbn, 0, 1])
        print("")
        print(f"Report written: {report_path}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())

