from __future__ import annotations

import argparse
import csv
import difflib
import re
from pathlib import Path
from typing import Any, Iterable, Optional


AR_MAP = str.maketrans(
    {
        "أ": "ا",
        "إ": "ا",
        "آ": "ا",
        "ٱ": "ا",
        "ى": "ي",
        "ئ": "ي",
    }
)

AR_DIACRITICS_RE = re.compile(r"[\u0610-\u061A\u064B-\u065F\u0670\u06D6-\u06ED\u0640]+")
NON_WORD_RE = re.compile(r"[^\w\s]+", flags=re.UNICODE)


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Match XLSX titles to local cover filenames (normalized Arabic)")
    p.add_argument("--xlsx", required=True, help="Path to Excel file")
    p.add_argument(
        "--covers-dir",
        default=str(Path(__file__).resolve().parents[1] / "storage" / "app" / "covers_by_title"),
        help="Directory containing exported covers named by normalized title",
    )
    p.add_argument("--sheet", default=None, help="Sheet name (default: active sheet)")
    p.add_argument("--title-col", type=int, default=0, help="1-based title column override (0 = auto-detect)")
    p.add_argument("--max-scan-rows", type=int, default=0, help="Max data rows to scan (0 = all)")
    p.add_argument("--sample-rows", type=int, default=400, help="Rows used for auto-detection scoring")
    p.add_argument("--threshold", type=float, default=0.98, help="Fuzzy match threshold (0-1), e.g. 0.98")
    p.add_argument(
        "--max-candidates",
        type=int,
        default=250,
        help="Max candidate covers to score per title (after token filtering)",
    )
    p.add_argument("--write-report", action="store_true", help="Write CSV report next to covers dir")
    return p.parse_args()


def norm_spaces(s: str) -> str:
    s = str(s or "")
    s = s.replace("\u00A0", " ")
    s = re.sub(r"\s+", " ", s, flags=re.UNICODE).strip()
    return s


def norm_title_key(val: Any) -> str:
    if val is None:
        return ""
    if isinstance(val, bool):
        return ""
    s = norm_spaces(str(val))
    if not s:
        return ""
    s = s.translate(AR_MAP)
    s = AR_DIACRITICS_RE.sub("", s)
    s = NON_WORD_RE.sub(" ", s)
    s = re.sub(r"\s+", " ", s, flags=re.UNICODE).strip()
    return s.casefold()


def iter_rows(ws, start_row: int = 1) -> Iterable[list[Any]]:
    for row in ws.iter_rows(min_row=start_row, values_only=True):
        yield list(row)

def tokenize_title_key(k: str) -> list[str]:
    # Keep only meaningful tokens (Arabic words, latin, numbers)
    toks = [t for t in (k or "").split(" ") if len(t) >= 2]
    return toks


def auto_detect_title_col(ws, sample_rows: int) -> tuple[Optional[int], str]:
    """
    Auto-detect title column in sheets without headers by scoring columns.
    Heuristic: titles tend to be string-heavy with higher uniqueness and longer average length.
    Returns (0-based col idx, method).
    """
    rows = []
    for i, row in enumerate(iter_rows(ws, start_row=1), start=1):
        rows.append(row)
        if i >= max(50, sample_rows):
            break

    if not rows:
        return None, "no_rows"

    max_cols = max(len(r) for r in rows)
    best = None
    best_score = -1.0

    for c in range(max_cols):
        vals = []
        for r in rows:
            v = r[c] if c < len(r) else None
            if v is None:
                continue
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                continue
            s = norm_spaces(str(v))
            if len(s) < 2:
                continue
            vals.append(s)

        if len(vals) < 50:
            continue

        keys = [norm_title_key(v) for v in vals]
        keys = [k for k in keys if k]
        if len(keys) < 50:
            continue

        uniq = len(set(keys))
        avg_len = sum(len(v) for v in vals) / max(1, len(vals))
        # Penalize columns that look like categories (very low uniqueness)
        score = (len(vals) * 1.0) + (uniq * 1.5) + (avg_len * 0.3)

        if score > best_score:
            best_score = score
            best = c

    if best is None:
        return None, "auto_detect_failed"

    return best, f"auto_detect(score={best_score:.1f})"


def scan_xlsx_titles(xlsx_path: Path, sheet: Optional[str], title_col_1based: int, max_scan_rows: int, sample_rows: int):
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as e:  # pragma: no cover
        raise SystemExit(f"openpyxl is required to read XLSX. Error: {e}")

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb.active

    if title_col_1based > 0:
        title_col = title_col_1based - 1
        method = "manual"
    else:
        title_col, method = auto_detect_title_col(ws, sample_rows=sample_rows)

    meta = {"sheet": ws.title, "title_col": title_col, "detection_method": method}
    if title_col is None:
        return set(), meta

    titles: set[str] = set()
    scanned = 0
    empty = 0

    for row in iter_rows(ws, start_row=1):
        if max_scan_rows > 0 and scanned >= max_scan_rows:
            break
        scanned += 1
        v = row[title_col] if title_col < len(row) else None
        k = norm_title_key(v)
        if not k:
            empty += 1
            continue
        titles.add(k)

    meta.update({"scanned_rows": scanned, "empty_title_rows": empty, "unique_titles": len(titles)})
    return titles, meta


def scan_cover_title_keys(covers_dir: Path) -> tuple[set[str], dict[str, Any]]:
    if not covers_dir.exists() or not covers_dir.is_dir():
        raise SystemExit(f"covers dir not found: {covers_dir}")

    keys: set[str] = set()
    total_files = 0
    for p in covers_dir.iterdir():
        if not p.is_file():
            continue
        if p.name.startswith("_"):
            continue
        total_files += 1
        base = p.stem
        # strip collision suffix
        base = re.sub(r"__id\d+$", "", base)
        k = norm_title_key(base.replace("-", " "))
        if k:
            keys.add(k)

    meta = {"covers_dir": str(covers_dir), "total_files": total_files, "unique_title_keys": len(keys)}
    return keys, meta


def build_token_index(keys: list[str]) -> dict[str, list[int]]:
    idx: dict[str, list[int]] = {}
    for i, k in enumerate(keys):
        for t in set(tokenize_title_key(k)):
            idx.setdefault(t, []).append(i)
    return idx


def best_fuzzy_match(
    query_key: str,
    cover_keys: list[str],
    cover_key_set: set[str],
    token_index: dict[str, list[int]],
    threshold: float,
    max_candidates: int,
) -> tuple[bool, str, float, str]:
    """
    Returns (matched, best_key, best_score, method)
    """
    if not query_key:
        return False, "", 0.0, "empty"

    if query_key in cover_key_set:
        return True, query_key, 1.0, "exact"

    toks = tokenize_title_key(query_key)
    if not toks:
        return False, "", 0.0, "no_tokens"

    # Candidate counts by shared tokens
    cand_counts: dict[int, int] = {}
    for t in set(toks):
        for i in token_index.get(t, []):
            cand_counts[i] = cand_counts.get(i, 0) + 1

    if not cand_counts:
        return False, "", 0.0, "no_candidates"

    # Take top candidates by token overlap (then shorter length difference)
    # to keep scoring fast.
    scored_cands = sorted(
        cand_counts.items(),
        key=lambda kv: (kv[1], -abs(len(cover_keys[kv[0]]) - len(query_key))),
        reverse=True,
    )
    cand_ids = [i for i, _ in scored_cands[: max(1, max_candidates)]]

    best_key = ""
    best_score = 0.0
    for i in cand_ids:
        ck = cover_keys[i]
        # Quick length gate: if very different, ratio won't reach 0.98 anyway.
        if threshold >= 0.95:
            if abs(len(ck) - len(query_key)) > max(3, int((1 - threshold) * max(len(ck), len(query_key)) * 2)):
                # still allow if it shares many tokens; we already filtered, but keep it fast
                pass
        s = difflib.SequenceMatcher(a=query_key, b=ck).ratio()
        if s > best_score:
            best_score = s
            best_key = ck
            if best_score >= 0.999:
                break

    matched = best_score >= threshold
    return matched, best_key, float(best_score), "fuzzy"


def main() -> int:
    args = parse_args()
    xlsx = Path(args.xlsx)
    covers_dir = Path(args.covers_dir)

    excel_titles, excel_meta = scan_xlsx_titles(
        xlsx_path=xlsx,
        sheet=args.sheet,
        title_col_1based=args.title_col,
        max_scan_rows=args.max_scan_rows,
        sample_rows=args.sample_rows,
    )
    cover_key_set, cover_meta = scan_cover_title_keys(covers_dir)
    cover_keys = sorted(cover_key_set)
    token_index = build_token_index(cover_keys)

    threshold = float(args.threshold)
    max_candidates = int(args.max_candidates)

    matched: set[str] = set()
    excel_only: set[str] = set()
    best_map: dict[str, tuple[str, float, str]] = {}

    for k in excel_titles:
        ok, best_k, score, method = best_fuzzy_match(
            query_key=k,
            cover_keys=cover_keys,
            cover_key_set=cover_key_set,
            token_index=token_index,
            threshold=threshold,
            max_candidates=max_candidates,
        )
        if ok:
            matched.add(k)
            best_map[k] = (best_k, score, method)
        else:
            excel_only.add(k)
            best_map[k] = (best_k, score, method)

    covers_only = cover_key_set  # will subtract matched covers below (optional)
    used_covers = {best_map[k][0] for k in matched if best_map[k][0]}
    covers_only = cover_key_set - used_covers

    print("XLSX:")
    print(f"  file: {xlsx}")
    print(f"  sheet: {excel_meta.get('sheet')}")
    print(f"  title_col: {excel_meta.get('title_col')}  detection: {excel_meta.get('detection_method')}")
    print(f"  scanned_rows: {excel_meta.get('scanned_rows')}")
    print(f"  unique_titles: {excel_meta.get('unique_titles')}")
    print("")
    print("COVERS:")
    print(f"  dir: {covers_dir}")
    print(f"  total_files: {cover_meta.get('total_files')}")
    print(f"  unique_title_keys: {cover_meta.get('unique_title_keys')}")
    print("")
    print("MATCHING (by normalized title, fuzzy):")
    print(f"  threshold: {threshold}")
    print(f"  matched_unique_titles: {len(matched)}")
    print(f"  excel_titles_without_cover: {len(excel_only)}")
    print(f"  cover_titles_not_in_excel: {len(covers_only)}")

    if args.write_report:
        report_path = covers_dir / "_title_match_report.csv"
        with report_path.open("w", encoding="utf-8", newline="") as f:
            w = csv.writer(f)
            w.writerow(["excel_title_key", "matched", "best_cover_title_key", "score", "method"])
            for k in sorted(excel_titles):
                best_k, score, method = best_map.get(k, ("", 0.0, ""))
                w.writerow([k, 1 if k in matched else 0, best_k, f"{score:.4f}", method])
        print("")
        print(f"Report written: {report_path}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())

