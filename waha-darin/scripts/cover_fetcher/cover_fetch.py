from __future__ import annotations

import argparse
import csv
import io
import mimetypes
import os
import sys
import time
import urllib.parse
from dataclasses import asdict
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout

from common import BookRow, ensure_dir, now_iso, norm_spaces, safe_slug

try:
    from PIL import Image
except Exception:  # pragma: no cover
    Image = None  # type: ignore


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser()
    p.add_argument("--xlsx", required=True)
    p.add_argument("--sheet", default="Sheet1")
    p.add_argument("--out-dir", required=True)
    p.add_argument("--mapping", required=True)
    p.add_argument("--limit", type=int, default=50)
    p.add_argument("--start-row", type=int, default=1)
    p.add_argument("--user-data-dir", default=None)
    p.add_argument("--slow-mo-ms", type=int, default=0)
    p.add_argument(
        "--headless",
        type=int,
        default=None,
        help="1=headless, 0=headed. Default: headless if no DISPLAY, else headed.",
    )
    p.add_argument(
        "--browser",
        default="chromium",
        choices=["chromium", "webkit", "firefox"],
        help="Browser engine to use. If chromium crashes on M1, use webkit.",
    )
    p.add_argument(
        "--manual-wait-secs",
        type=int,
        default=180,
        help="How long to wait for Google Images thumbnails (time for manual consent/captcha).",
    )
    p.add_argument(
        "--max-results-to-try",
        type=int,
        default=12,
        help="How many top Google Images results to try before giving up.",
    )
    p.add_argument(
        "--first-n",
        dest="max_results_to_try",
        type=int,
        default=12,
        help="Alias for --max-results-to-try (how many first results to consider).",
    )
    return p.parse_args()


def is_missing_cover(has_image: str, cover_name: str) -> bool:
    hi = norm_spaces(has_image)
    cn = norm_spaces(cover_name)
    # In our sheets, has_image is usually "0"/"1" as string
    if hi == "1" and cn:
        return False
    return True


def iter_missing_rows(xlsx: Path, sheet: str, start_row: int) -> Iterable[BookRow]:
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        raise SystemExit(f"Sheet not found: {sheet}. Sheets: {wb.sheetnames}")
    ws = wb[sheet]

    # DB-export style columns:
    # C=title, D=author, E=publisher, G=ISBN, AA=has_image, AB=cover_name
    for r in range(start_row, ws.max_row + 1):
        title = norm_spaces(ws[f"C{r}"].value or "")
        if not title:
            continue
        author = norm_spaces(ws[f"D{r}"].value or "")
        publisher = norm_spaces(ws[f"E{r}"].value or "")
        isbn = norm_spaces(ws[f"G{r}"].value or "")
        has_image = norm_spaces(ws[f"AA{r}"].value or "")
        cover_name = norm_spaces(ws[f"AB{r}"].value or "")

        if is_missing_cover(has_image, cover_name):
            yield BookRow(row=r, title=title, author=author, publisher=publisher, isbn=isbn)


def read_existing_mapping_rows(mapping_path: Path, out_dir: Path) -> set[int]:
    """
    Returns row numbers that should be skipped.

    Rules:
    - accepted/deleted => skip
    - downloaded => skip ONLY if the file still exists on disk
    - error => do not skip (allow retry)
    """
    if not mapping_path.exists():
        return set()

    last: dict[int, dict[str, str]] = {}
    with mapping_path.open("r", encoding="utf-8", newline="") as f:
        r = csv.DictReader(f, delimiter="\t")
        for row in r:
            try:
                row_num = int(row.get("row") or 0)
            except ValueError:
                continue
            if row_num <= 0:
                continue
            last[row_num] = {
                "status": (row.get("status") or "").strip().lower(),
                "saved_filename": (row.get("saved_filename") or "").strip(),
            }

    done: set[int] = set()
    for row_num, info in last.items():
        st = info.get("status", "")
        fn = info.get("saved_filename", "")
        if st in {"accepted", "deleted"}:
            done.add(row_num)
        elif st == "downloaded":
            if fn and (out_dir / fn).exists():
                done.add(row_num)
    return done


def ensure_mapping_header(mapping_path: Path) -> None:
    if mapping_path.exists():
        return
    ensure_dir(mapping_path.parent)
    with mapping_path.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f, delimiter="\t")
        w.writerow(
            [
                "row",
                "title",
                "author",
                "publisher",
                "isbn",
                "query",
                "image_url",
                "saved_filename",
                "status",
                "downloaded_at",
                "reviewed_at",
            ]
        )


def build_query(b: BookRow) -> str:
    """
    User request: search by book name + author.
    We keep "غلاف كتاب" to bias toward covers.
    """
    title = norm_spaces(b.title)
    author = norm_spaces(b.author)
    publisher = norm_spaces(b.publisher)

    if author and publisher:
        return f"\"{title}\" \"{author}\" \"{publisher}\" غلاف كتاب"
    if author:
        return f"\"{title}\" \"{author}\" غلاف كتاب"
    if publisher:
        return f"\"{title}\" \"{publisher}\" غلاف كتاب"
    return f"\"{title}\" غلاف كتاب"


def thumb_url(img) -> str:
    src = (img.get_attribute("src") or "").strip()
    dsrc = (img.get_attribute("data-src") or "").strip()
    return src or dsrc or ""


def guess_ext(content_type: str | None, url: str) -> str:
    ct = (content_type or "").split(";")[0].strip().lower()
    if ct:
        ext = mimetypes.guess_extension(ct) or ""
        if ext:
            return ext.lstrip(".")
    # fallback from URL
    path = urllib.parse.urlparse(url).path
    ext = os.path.splitext(path)[1].lower().lstrip(".")
    if ext in {"jpg", "jpeg", "png", "webp", "gif"}:
        return ext
    return "jpg"


def unique_filename(out_dir: Path, base: str, ext: str) -> str:
    base = base.strip()
    ext = ext.strip(".").lower()
    cand = f"{base}.{ext}"
    if not (out_dir / cand).exists():
        return cand
    n = 2
    while True:
        cand = f"{base}-{n}.{ext}"
        if not (out_dir / cand).exists():
            return cand
        n += 1


def google_images_url(query: str) -> str:
    q = urllib.parse.quote_plus(query)
    return f"https://www.google.com/search?tbm=isch&q={q}"


def extract_imgurl_from_href(href: str) -> str | None:
    """
    Google images result links often look like:
    https://www.google.com/imgres?imgurl=<ENCODED_URL>&imgrefurl=...
    """
    href = (href or "").strip()
    if not href:
        return None
    try:
        u = urllib.parse.urlparse(href)
        qs = urllib.parse.parse_qs(u.query)
        for key in ("imgurl", "mediaurl"):
            if key in qs and qs[key]:
                return urllib.parse.unquote(qs[key][0])
    except Exception:
        return None
    return None


def closest_anchor_href(img) -> str:
    try:
        return img.evaluate("el => el.closest('a')?.href || ''") or ""
    except Exception:
        return ""


def validate_image_bytes(body: bytes) -> tuple[int, int]:
    """
    Reject tiny placeholders/logos by checking dimensions when Pillow is available.
    Fallback: basic size threshold.
    """
    if Image is None:
        if len(body) < 30_000:
            raise RuntimeError(f"Downloaded file too small ({len(body)} bytes), rejecting")
        return (0, 0)

    try:
        with Image.open(io.BytesIO(body)) as im:
            w, h = im.size
            if w < 220 or h < 220:
                raise RuntimeError(f"Image too small ({w}x{h}), rejecting")
            ratio = w / max(h, 1)
            # Reject very wide/short images (often search bars/banners)
            if ratio > 2.0 or ratio < 0.35:
                raise RuntimeError(f"Bad aspect ratio ({w}x{h}), rejecting")
            return (int(w), int(h))
    except RuntimeError:
        raise
    except Exception as e:
        raise RuntimeError(f"Unable to parse image bytes: {e}")

def find_first_result_thumbnail(page, manual_wait_secs: int):
    """
    Return a Playwright element handle for the first real Google Images result thumbnail.
    We avoid Google logos by requiring encrypted thumbnail URLs.
    """
    deadline = time.time() + max(5, manual_wait_secs)
    while time.time() < deadline:
        imgs = page.query_selector_all("img")
        for img in imgs:
            src = (img.get_attribute("src") or "").strip()
            dsrc = (img.get_attribute("data-src") or "").strip()
            cand = src or dsrc
            if not cand:
                continue
            if "encrypted-tbn" in cand or "gstatic.com/images?q=tbn" in cand:
                return img
        page.wait_for_timeout(250)
    return None


def list_result_thumbnails(page, manual_wait_secs: int):
    """
    Collect top Google Images result thumbnail <img> elements (encrypted-tbn*).
    We return element handles so the caller can click and extract full-size URLs.
    """
    deadline = time.time() + max(5, manual_wait_secs)
    seen: set[str] = set()
    while time.time() < deadline:
        out = []
        imgs = page.query_selector_all("img")
        for img in imgs:
            tsrc = thumb_url(img)
            if not tsrc:
                continue
            if not ("encrypted-tbn" in tsrc or "gstatic.com/images?q=tbn" in tsrc):
                continue
            if tsrc in seen:
                continue
            seen.add(tsrc)
            out.append(img)
        if out:
            return out
        page.wait_for_timeout(250)
    return []


def extract_first_full_image_url(page) -> str | None:
    """
    Clicked image viewer can contain multiple <img>. We pick first http(s) non-data.
    Prefer common Google viewer selectors, then fall back to scanning all imgs.
    """
    preferred_selectors = [
        "img.iPVvYb",  # common full-size image class
        "img.sFlh5c",  # sometimes used
    ]

    for _ in range(40):
        for sel in preferred_selectors:
            imgs = page.query_selector_all(sel)
            for img in imgs:
                src = (img.get_attribute("src") or "").strip()
                if src.startswith(("http://", "https://")):
                    return src

        # Fallback: scan all images, skip tiny icons and data URIs
        imgs = page.query_selector_all("img")
        for img in imgs:
            src = (img.get_attribute("src") or "").strip()
            if not src or src.startswith("data:"):
                continue
            if not src.startswith(("http://", "https://")):
                continue
            w = img.get_attribute("width") or ""
            h = img.get_attribute("height") or ""
            try:
                if w and h and int(w) < 140 and int(h) < 140:
                    continue
            except ValueError:
                pass
            return src

        page.wait_for_timeout(250)
    return None


def try_click_consent(page) -> None:
    # Best-effort; if not found, ignore.
    candidates = [
        "button:has-text('I agree')",
        "button:has-text('Accept all')",
        "button:has-text('Accept')",
        "button:has-text('Ich stimme zu')",
        "button:has-text('Alle akzeptieren')",
    ]
    for sel in candidates:
        try:
            btn = page.query_selector(sel)
            if btn:
                btn.click(timeout=1500)
                return
        except Exception:
            pass


def main() -> int:
    args = parse_args()
    xlsx = Path(args.xlsx)
    out_dir = Path(args.out_dir)
    mapping_path = Path(args.mapping)

    print("cover_fetch starting…", flush=True)
    t0 = time.time()

    ensure_dir(out_dir)
    ensure_mapping_header(mapping_path)

    already_done = read_existing_mapping_rows(mapping_path, out_dir)
    print("Scanning XLSX for missing-cover rows…", flush=True)

    to_process: list[BookRow] = []
    for b in iter_missing_rows(xlsx, args.sheet, args.start_row):
        if b.row in already_done:
            continue
        to_process.append(b)
        if len(to_process) >= max(0, args.limit):
            break

    if not to_process:
        print("No missing-cover rows to process (or all already in mapping).", flush=True)
        return 0

    print(f"Will process now: {len(to_process)}", flush=True)
    print(f"Download folder: {out_dir}", flush=True)
    print(f"Mapping TSV: {mapping_path}", flush=True)
    print(f"XLSX scan time: {time.time() - t0:.2f}s", flush=True)

    user_data_dir = args.user_data_dir or str(out_dir / ".pw-profile")
    if args.headless is None:
        # In many automation environments there is no GUI; default to headless then.
        args.headless = 0 if os.environ.get("DISPLAY") else 1
    headless = bool(int(args.headless))
    print(f"Browser mode: {'headless' if headless else 'headed'}", flush=True)

    with sync_playwright() as p:
        def launch(engine: str):
            browser_type = getattr(p, engine)
            return browser_type.launch_persistent_context(
                user_data_dir=user_data_dir,
                headless=headless,
                slow_mo=args.slow_mo_ms,
                locale="en-US",
            )

        print(f"Launching browser context ({args.browser})…", flush=True)
        try:
            context = launch(args.browser)
        except Exception as e:
            if args.browser != "chromium":
                raise
            # Chromium sometimes crashes on Apple Silicon if wrong build is present; fallback to WebKit.
            print(f"Chromium launch failed ({e}). Falling back to WebKit…", flush=True)
            context = launch("webkit")
        page = context.new_page()

        # Warm up
        print("Opening google.com…", flush=True)
        page.goto("https://www.google.com", wait_until="domcontentloaded", timeout=60000)

        for idx, book in enumerate(to_process, start=1):
            query = build_query(book)
            url = google_images_url(query)
            print(f"[{idx}/{len(to_process)}] row={book.row} query={query}")

            image_url = ""
            saved_filename = ""
            status = "downloaded"
            downloaded_at = now_iso()

            try:
                page.goto(url, wait_until="domcontentloaded", timeout=60000)
                try_click_consent(page)

                # Wait for a real image-result thumbnail (encrypted-tbn*)
                print(
                    f"  Waiting for image results (up to {args.manual_wait_secs}s). "
                    f"If you see consent/captcha, solve it in the browser…",
                    flush=True,
                )
                thumbs = list_result_thumbnails(page, args.manual_wait_secs)
                if not thumbs:
                    raise RuntimeError("No Google Images result thumbnails found (blocked/consent?)")
                last_err: Exception | None = None
                best_area = -1
                best_body: bytes | None = None
                best_ct = ""
                best_url = ""
                best_wh = (0, 0)
                for j, thumb in enumerate(thumbs[: max(1, args.max_results_to_try)], start=1):
                    # Prefer original URL from the result link, otherwise click and extract full-size from viewer.
                    candidate_url = ""
                    try:
                        href = closest_anchor_href(thumb)
                        candidate_url = (extract_imgurl_from_href(href) or "").strip()
                    except Exception:
                        candidate_url = ""

                    if not candidate_url:
                        try:
                            thumb.click(timeout=10_000)
                            time.sleep(0.6)
                            candidate_url = (extract_first_full_image_url(page) or "").strip()
                        except Exception:
                            candidate_url = ""

                    if not candidate_url:
                        # Never download tiny thumbnails as "covers"
                        last_err = RuntimeError("Could not extract full-size image URL")
                        continue

                    if "fonts.gstatic.com" in candidate_url or "productlogos" in candidate_url:
                        last_err = RuntimeError("Google asset URL")
                        continue
                    try:
                        resp = context.request.get(candidate_url, timeout=60000)
                        if not resp.ok:
                            raise RuntimeError(f"HTTP {resp.status}")
                        ct = (resp.headers.get("content-type") or "").lower()
                        if ct and "image" not in ct:
                            raise RuntimeError(f"not image content-type: {ct}")
                        if "svg" in ct:
                            raise RuntimeError("svg")
                        body = resp.body()
                        w, h = validate_image_bytes(body)
                        area = (w * h) if (w and h) else len(body)
                        if w and h:
                            print(f"  candidate {j}: {w}x{h}", flush=True)
                        if area > best_area:
                            best_area = area
                            best_body = body
                            best_ct = ct
                            best_url = candidate_url
                            best_wh = (w, h)
                        last_err = None
                    except Exception as e:
                        last_err = e
                    # Best-effort: close viewer overlay so next click works reliably.
                    try:
                        page.keyboard.press("Escape")
                    except Exception:
                        pass

                if best_body is None:
                    raise RuntimeError(f"All candidates rejected. Last error: {last_err}")

                ext = guess_ext(best_ct, best_url)
                base = f"{safe_slug(book.title, 90)}__{safe_slug(book.publisher, 60)}__r{book.row}"
                saved_filename = unique_filename(out_dir, base, ext)
                (out_dir / saved_filename).write_bytes(best_body)
                image_url = best_url
                if best_wh != (0, 0):
                    print(f"  Picked best image: {best_wh[0]}x{best_wh[1]}", flush=True)
            except PWTimeout as e:
                status = "error"
                print(f"  ERROR timeout: {e}")
            except Exception as e:
                status = "error"
                print(f"  ERROR: {e}")

            # Append to mapping TSV
            with mapping_path.open("a", encoding="utf-8", newline="") as f:
                w = csv.writer(f, delimiter="\t")
                w.writerow(
                    [
                        book.row,
                        book.title,
                        book.author,
                        book.publisher,
                        book.isbn,
                        query,
                        image_url,
                        saved_filename,
                        status,
                        downloaded_at,
                        "",
                    ]
                )

        context.close()

    print("Done.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

