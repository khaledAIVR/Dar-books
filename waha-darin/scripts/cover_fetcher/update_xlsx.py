from __future__ import annotations

import argparse
import csv
import shutil
from pathlib import Path

from openpyxl import load_workbook

from common import norm_spaces


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser()
    p.add_argument("--xlsx-in", required=True)
    p.add_argument("--xlsx-out", required=True)
    p.add_argument("--sheet", default="Sheet1")
    p.add_argument("--mapping", required=True)
    return p.parse_args()


def read_accepted(mapping: Path) -> dict[int, str]:
    """
    Returns row_number => saved_filename for accepted rows.
    """
    out: dict[int, str] = {}
    with mapping.open("r", encoding="utf-8", newline="") as f:
        r = csv.DictReader(f, delimiter="\t")
        for d in r:
            status = norm_spaces(d.get("status") or "")
            if status != "accepted":
                continue
            try:
                row_num = int(d.get("row") or 0)
            except ValueError:
                continue
            fn = norm_spaces(d.get("saved_filename") or "")
            if row_num > 0 and fn:
                out[row_num] = fn
    return out


def main() -> int:
    args = parse_args()
    xlsx_in = Path(args.xlsx_in)
    xlsx_out = Path(args.xlsx_out)
    mapping = Path(args.mapping)

    if not xlsx_in.exists():
        raise SystemExit(f"Input XLSX not found: {xlsx_in}")
    if not mapping.exists():
        raise SystemExit(f"Mapping TSV not found: {mapping}")

    accepted = read_accepted(mapping)
    if not accepted:
        print("No accepted covers found in mapping. Nothing to apply.")
        return 0

    # Copy input to output first (so we preserve as much as possible)
    xlsx_out.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(xlsx_in, xlsx_out)

    wb = load_workbook(xlsx_out)
    if args.sheet not in wb.sheetnames:
        raise SystemExit(f"Sheet not found: {args.sheet}. Sheets: {wb.sheetnames}")
    ws = wb[args.sheet]

    updated = 0
    for row_num, filename in accepted.items():
        # AA = has_image, AB = cover_name
        ws[f"AA{row_num}"].value = "1"
        ws[f"AB{row_num}"].value = filename
        updated += 1

    wb.save(xlsx_out)
    print(f"Applied accepted covers: {updated}")
    print(f"Wrote: {xlsx_out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

