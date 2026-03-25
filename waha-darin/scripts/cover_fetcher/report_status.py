from __future__ import annotations

import argparse
import csv
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

from common import norm_spaces


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser()
    p.add_argument("--xlsx", required=True)
    p.add_argument("--sheet", default="Sheet1")
    p.add_argument("--mapping", required=True)
    p.add_argument("--out-dir", required=True, help="Folder that contains downloaded covers")
    return p.parse_args()


def read_mapping(mapping: Path) -> list[dict[str, str]]:
    if not mapping.exists():
        return []
    with mapping.open("r", encoding="utf-8", newline="") as f:
        r = csv.DictReader(f, delimiter="\t")
        return [{k: norm_spaces(v or "") for k, v in row.items()} for row in r]


def count_missing_in_xlsx(xlsx: Path, sheet: str) -> tuple[int, int]:
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb[sheet]
    total_with_title = 0
    missing = 0
    for r in range(1, ws.max_row + 1):
        title = norm_spaces(ws[f"C{r}"].value or "")
        if not title:
            continue
        total_with_title += 1
        has_image = norm_spaces(ws[f"AA{r}"].value or "")
        cover_name = norm_spaces(ws[f"AB{r}"].value or "")
        if not (has_image == "1" and cover_name):
            missing += 1
    return total_with_title, missing


def main() -> int:
    args = parse_args()
    xlsx = Path(args.xlsx)
    mapping = Path(args.mapping)
    out_dir = Path(args.out_dir)

    rows = read_mapping(mapping)
    c = Counter((r.get("status") or "").strip() for r in rows)

    files_on_disk = 0
    if out_dir.exists():
        for p in out_dir.iterdir():
            if p.is_file() and not p.name.startswith(".") and not p.name.endswith(".tsv"):
                files_on_disk += 1

    total, missing = count_missing_in_xlsx(xlsx, args.sheet)

    print(f"XLSX: {xlsx}")
    print(f"Sheet: {args.sheet}")
    print(f"Rows with title: {total}")
    print(f"Rows missing cover (AA/AB not set): {missing}")
    print("")
    print(f"Mapping TSV: {mapping}")
    print("Status counts:")
    for k in sorted(c.keys()):
        print(f"  {k or '(blank)'}: {c[k]}")
    print("")
    print(f"Downloaded files on disk: {files_on_disk} ({out_dir})")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

